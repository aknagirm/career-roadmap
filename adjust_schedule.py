"""
adjust_schedule.py
──────────────────
Gap Day Handler for the 7-Month Interview Prep Schedule.

How it works:
1. Reads the existing Excel schedule.
2. Finds every row where "Done" is blank AND the date is in the past
   — these are treated as gap days (days you couldn't study).
3. Shifts all future unfinished topic rows forward by the number of gap days found,
   inserting actual calendar dates in sequence.
4. Overwrites the Excel file with the updated schedule.

Usage:
    python adjust_schedule.py

Notes:
- Rows already marked "Completed" or "In Progress" are never moved.
- Gap detection only applies to past dates (before today) with no "Done" value.
- Saturday and Sunday rows (Mock / Project) are treated the same as weekdays —
  they shift forward too.
- The script preserves all formatting, colors, and the "Done" dropdown.
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = "6_month_interview_prep_schedule.xlsx"

THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FOCUS_COLORS = {
    "JavaScript":                "FFF2CC",
    "Angular":                   "D9EAD3",
    "TypeScript":                "E2EFDA",
    "DSA":                       "FCE5CD",
    "System Design":             "CFE2F3",
    "AI Engineering":            "EAD1DC",
    "Backend":                   "D9D2E9",
    "LeetCode":                  "F4CCCC",
    "Behavioral":                "FFF2CC",
    "Testing":                   "D0E4F7",
    "Mock Interview + Revision": "FFE0B2",
    "Build / Read / Explore":    "E8F5E9",
}

MONTH_ROW_COLOR = "BDD7EE"
HEADER_COLOR    = "1F4E79"


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def load_schedule(ws):
    """
    Read all data rows from the schedule sheet.
    Returns:
        headers   : list of column header names
        data_rows : list of dicts, one per actual schedule row (skips month separators)
    """
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    data_rows = []

    for row_idx in range(2, ws.max_row + 1):
        row_vals = {headers[c]: ws.cell(row=row_idx, column=c + 1).value
                    for c in range(len(headers))}

        # Skip month separator rows (merged cells, no "Date" value that looks like a date)
        date_val = ws.cell(row=row_idx, column=3).value  # "Date" is column 3
        if date_val is None or str(date_val).startswith("━"):
            continue

        notes_col = headers.index("Notes") + 1 if "Notes" in headers else None
        done_col  = headers.index("Done")  + 1 if "Done"  in headers else None

        data_rows.append({
            "week":    ws.cell(row=row_idx, column=1).value,
            "month":   ws.cell(row=row_idx, column=2).value,
            "date_str":ws.cell(row=row_idx, column=3).value,
            "day":     ws.cell(row=row_idx, column=4).value,
            "focus":   ws.cell(row=row_idx, column=5).value,
            "topic":   ws.cell(row=row_idx, column=6).value,
            "done":    ws.cell(row=row_idx, column=7).value,
            "notes":   ws.cell(row=row_idx, column=8).value,
        })

    return data_rows


def parse_date(date_str: str):
    """Parse date string in 'DD Mon YYYY' format."""
    try:
        return date(*[int(x) for x in
                      __import__("datetime").datetime.strptime(date_str, "%d %b %Y").timetuple()[:3]])
    except Exception:
        return None


def adjust(data_rows: list) -> list:
    """
    Core adjustment logic:
    - Find past rows with no "Done" value → gap days
    - Count gap days
    - Shift all future unfinished rows forward by gap_count days
    - Reassign dates, day names, week numbers, month labels
    """
    today = date.today()
    gap_count = 0

    for row in data_rows:
        d = parse_date(row["date_str"])
        done = row["done"]
        if d and d < today and not done:
            gap_count += 1

    if gap_count == 0:
        print("✅ No gap days found. Schedule is up to date.")
        return data_rows

    print(f"📋 Found {gap_count} gap day(s). Shifting future unfinished rows forward...")

    # Find the first future unfinished row
    future_rows = []
    past_done_rows = []

    for row in data_rows:
        d = parse_date(row["date_str"])
        done = row["done"]
        if done in ("Completed", "In Progress") or (d and d < today and not done):
            past_done_rows.append(row)
        else:
            future_rows.append(row)

    # Assign new dates to future rows starting from today + 1 day gap
    if not future_rows:
        print("✅ No future rows to shift.")
        return data_rows

    # Find what date the first future row currently has
    first_future_date = parse_date(future_rows[0]["date_str"])
    new_start = first_future_date + timedelta(days=gap_count) if first_future_date else today + timedelta(days=1)

    current = new_start
    for row in future_rows:
        row["date_str"] = current.strftime("%d %b %Y")
        row["day"]      = current.strftime("%A")
        # Recalculate week and month relative to original start date
        # Keep existing week/month as approximate (full recalc would need START_DATE)
        current += timedelta(days=1)

    print(f"   First shifted date: {new_start.strftime('%d %b %Y')}")
    print(f"   Last shifted date:  {current.strftime('%d %b %Y')}")

    return past_done_rows + future_rows


def write_adjusted_excel(rows: list):
    """Rewrite the Excel file with adjusted rows, preserving formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "📅 7-Month Schedule"

    headers    = ["Week", "Month", "Date", "Day", "Focus Area", "Today's Topic / Task", "Done", "Notes"]
    col_widths = [7,      10,      14,     12,     22,           60,                     14,     30]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = make_fill(HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Done dropdown
    dv = DataValidation(
        type="list",
        formula1='"Completed,In Progress"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid entry",
        error='Please select "Completed" or "In Progress"',
    )
    dv.sqref = f"G2:G{len(rows) + 200}"
    ws.add_data_validation(dv)

    current_month = None
    excel_row = 2

    for r in rows:
        month_label = r.get("month", "")

        # Month separator
        if month_label != current_month:
            current_month = month_label
            sep_cell = ws.cell(row=excel_row, column=1,
                               value=f"━━━  {current_month}  ━━━")
            sep_cell.font      = Font(bold=True, size=11, color="1F4E79")
            sep_cell.fill      = make_fill(MONTH_ROW_COLOR)
            sep_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=len(headers))
            ws.row_dimensions[excel_row].height = 20
            excel_row += 1

        focus_color = FOCUS_COLORS.get(r.get("focus", ""), "FFFFFF")
        fill = make_fill(focus_color)

        values = [
            r.get("week"),
            r.get("month"),
            r.get("date_str"),
            r.get("day"),
            r.get("focus"),
            r.get("topic"),
            r.get("done"),
            r.get("notes"),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left" if col >= 5 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.font = Font(bold=(col == 5), size=9)

        ws.row_dimensions[excel_row].height = 28
        excel_row += 1

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Adjusted schedule saved: {OUTPUT_FILE}")
    print("   Open it in Excel or Google Sheets to see the updated dates.\n")


if __name__ == "__main__":
    print(f"\n📂 Loading schedule from {OUTPUT_FILE}...")

    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE)
    except FileNotFoundError:
        print(f"❌ File not found: {OUTPUT_FILE}")
        print("   Run generate_schedule.py first to create the schedule.\n")
        exit(1)

    ws = wb["📅 7-Month Schedule"]
    data_rows = load_schedule(ws)

    print(f"   Loaded {len(data_rows)} schedule rows.")

    adjusted_rows = adjust(data_rows)
    write_adjusted_excel(adjusted_rows)
