"""
30-Week Interview Preparation Schedule Generator
Architecture: Brain-learning model (deep focus per phase + continuous practice)

Learning Model (3 Layers):
  Layer 1 - Foundation: Deep focus on one topic cluster for 2-6 weeks
  Layer 2 - Continuous: DSA (3x/week), LeetCode (2x/week), Behavioral (1x/week)
  Layer 3 - Long-term: AI, Backend, System Design sprinkled 1-2x/week

Weekly Pattern (fixed):
  Mon  - Core topic subtopic A                        (3 hrs)
  Tue  - Core topic subtopic B                        (3 hrs)
  Wed  - Core topic subtopic C  +  DSA problem        (3 hrs)
  Thu  - Core topic subtopic D  +  Layer3 topic       (3 hrs)
  Fri  - Core topic coding Qs  +  LeetCode            (3 hrs)
  Sat  - Mock Interview + Revision + Project          (5 hrs)
  Sun  - Reading + Portfolio + Rest                   (2 hrs)

Phase Order (30 weeks):
  Phase 1  Weeks  1-4  : JavaScript
  Phase 2  Weeks  5-6  : TypeScript
  Phase 3  Weeks  7-12 : Angular (Components→RxJS→NgRx→SSR→Testing)
  Phase 4  Weeks 13-16 : Frontend System Design
  Phase 5  Weeks 17-21 : Backend + HTTP + Security
  Phase 6  Weeks 22-26 : AI Engineering
  Phase 7  Weeks 27-30 : Mock Loops + Revision + Apply

Excel Sheets:
  1. Dashboard         - live stats
  2. Master Roadmap    - all lessons ordered
  3. Daily Planner     - day-by-day (main tracker)
  4. Revision Tracker  - spaced repetition (1d/7d/30d/mastered)
  5. Mock Interviews   - log with scores
  6. Behavioral Stories- STAR story tracker
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
START_DATE  = date(2026, 7, 21)
TOTAL_WEEKS = 30
OUTPUT_FILE = "interview_prep_tracker.xlsx"

# ─────────────────────────────────────────────
# PHASE DEFINITIONS
# Each phase = list of weekly "focus clusters"
# Each cluster = (week_label, core_subtopics[Mon-Fri core parts], layer3_topic)
# ─────────────────────────────────────────────

# Layer 3 rotation (sprinkled Thu slot, cycling through these)
LAYER3_ROTATION = [
    "System Design: Browser Rendering & Critical Path",
    "AI: LLM Fundamentals & Prompt Engineering",
    "System Design: Design an E-commerce Frontend",
    "AI: OpenAI API & Function Calling",
    "System Design: Design a Chat Application",
    "AI: RAG Architecture & Chunking",
    "System Design: Design YouTube Frontend",
    "AI: LangChain – chains, agents, memory",
    "System Design: Authentication (OAuth, JWT)",
    "AI: MCP – Model Context Protocol deep dive",
    "System Design: Design Google Docs (collaborative)",
    "AI: LangGraph – stateful agentic workflows",
    "System Design: Micro Frontends architecture",
    "AI: Vector Databases & Embeddings",
    "System Design: Design a Notification System",
    "AI: AI Agents – ReAct pattern",
    "System Design: Rate Limiter & Design Token System",
    "AI: AI evaluation & hallucination mitigation",
    "System Design: File Upload & Design System",
    "AI: Streaming (SSE), Ollama, local models",
]

# DSA rotation (Wed slot + standalone Fri DSA)
DSA_TOPICS = [
    "Arrays – Two Sum, Max Subarray (implement)",
    "Arrays – Rotation, Dutch National Flag",
    "Strings – Anagram, Palindrome (implement)",
    "Strings – Sliding Window pattern",
    "Hash Maps – Frequency counter pattern",
    "Hash Maps – Two Pointer + HashMap combo",
    "Linked List – Reverse, cycle detection (implement)",
    "Stack – Valid parentheses, min stack (implement)",
    "Queue – BFS traversal (implement)",
    "Binary Search – Classic & rotated array",
    "Trees – Inorder, preorder, postorder (implement)",
    "Trees – Level order BFS (implement)",
    "Trees – Height, diameter, LCA",
    "Graphs – BFS & DFS (implement)",
    "Graphs – Cycle detection",
    "Sliding Window – Max sum subarray",
    "Two Pointer – Sorted arrays",
    "Heap – Top K elements",
    "DP – Fibonacci, coin change",
    "DP – Knapsack basics",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
    "Timed Practice – 2 mediums in 45 min",
]

# LeetCode rotation (Fri slot)
LEETCODE_TOPICS = [
    "2 Mediums: Arrays/Strings",
    "2 Mediums: HashMap/Two Pointer",
    "2 Mediums: Trees/BFS/DFS",
    "2 Mediums: Binary Search",
    "2 Mediums: Sliding Window",
    "2 Mediums: Stack/Queue",
    "2 Mediums: Graphs",
    "2 Mediums: DP basics",
    "2 Mediums: Mixed review",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Timed mock – 3 problems in 60 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
    "Hard: 1 problem in 45 min",
]

# Behavioral rotation (sprinkled 1x/week – Fri secondary)
BEHAVIORAL_TOPICS = [
    "STAR: AngularJS→Angular 12 migration",
    "STAR: Leading 25-engineer team",
    "STAR: TypeScript standards adoption",
    "STAR: AI workflow automation at Boeing",
    "STAR: Mentoring junior engineers",
    "STAR: Handling technical disagreement",
    "STAR: Production issue / incident",
    "STAR: Code review process",
    "STAR: Work estimation & planning",
    "STAR: Technical decision you influenced",
    "Practice: Why are you leaving Boeing?",
    "Practice: 60-second personal pitch",
    "Practice: Salary negotiation talking points",
    "Practice: Questions to ask the interviewer",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
    "Full behavioral mock round (record yourself)",
]

# Mock interview topics (Sat)
MOCK_TOPICS = [
    "Mock: Angular lifecycle & Change Detection Q&A",
    "Mock: JavaScript async & closures Q&A",
    "Mock: TypeScript generics & mapped types",
    "Mock: Angular RxJS patterns Q&A",
    "Mock: Angular NgRx deep dive Q&A",
    "Mock: System Design – Chat App (60 min)",
    "Mock: System Design – E-commerce (60 min)",
    "Mock: JavaScript advanced coding round",
    "Mock: Behavioral round (45 min)",
    "Mock: Full loop – Frontend technical round",
    "Mock: Full loop – System Design round",
    "Mock: Full loop – Behavioral round",
    "Mock: Backend Node.js + REST Q&A",
    "Mock: Security & HTTP Q&A",
    "Mock: AI Engineering Q&A (45 min)",
    "Mock: Mixed technical round (60 min)",
    "Mock: Full loop simulation – Day 1",
    "Mock: Full loop simulation – Day 2",
    "Mock: Final weak-area targeted round",
    "Mock: Final full loop simulation",
    "Mock: Company-specific prep (Atlassian/Adobe)",
    "Mock: Company-specific prep (Flipkart/Razorpay)",
    "Mock: Company-specific prep (Salesforce/ServiceNow)",
    "Mock: Final confidence round",
    "Mock: Revision – all weak areas",
    "Mock: Revision – all weak areas",
    "Mock: Revision – all weak areas",
    "Mock: Revision – all weak areas",
    "Mock: Revision – all weak areas",
    "Mock: Revision – all weak areas",
]

# Sunday project/reading topics
SUNDAY_TOPICS = [
    "Project: convopro_chatgpt – plan streaming feature",
    "Read: JavaScript.info – advanced chapters",
    "Project: Add streaming response to chatgpt app",
    "Read: Angular blog – Signals deep dive",
    "Project: Implement RAG in chatgpt app",
    "Read: LangChain docs – agents",
    "Project: Add conversation history with MongoDB",
    "Read: Engineering blog – system design at scale",
    "Project: Deploy app to cloud (Vercel/Render)",
    "Read: Frontend performance case studies",
    "Project: Write LinkedIn post about AI project",
    "Read: Staff Engineer book / blog chapter",
    "Project: Polish GitHub README for portfolio",
    "Read: Watch AngularConnect / JSConf talk",
    "Project: Build Design Token demo in Angular",
    "Read: Micro Frontends case study",
    "Project: Explore LangGraph – build a flow",
    "Read: AI evaluation frameworks",
    "Project: AI Agent with tool calling (mini project)",
    "Read: Engineering culture blogs (Netflix, Uber)",
    "Project: Free build / explore new tool",
    "Read: Review all notes from the past months",
    "Project: Final portfolio cleanup & README polish",
    "Read: Watch a mock interview on YouTube",
    "Project: Write LinkedIn post – career journey",
    "Free: Rest day – light reading only",
    "Free: Rest day – light reading only",
    "Free: Rest day – light reading only",
    "Free: Rest day – light reading only",
    "Free: Rest day – light reading only",
]

# ─────────────────────────────────────────────
# CORE PHASE CURRICULUM
# Ordered list of (phase, week_in_phase, subtopics[5]) for Mon-Fri core slots
# subtopics[0]=Mon, [1]=Tue, [2]=Wed-core, [3]=Thu-core, [4]=Fri-core
# ─────────────────────────────────────────────

PHASE_WEEKS = [
    # ── PHASE 1: JavaScript (Weeks 1–4) ──────────────────────────────────────
    {
        "phase": "JavaScript", "week_label": "JS Week 1: Execution Context & Scope",
        "subtopics": [
            "Execution Context, Call Stack, Variable Environment",
            "Scope Chain, Lexical Scope, Hoisting (var/let/const/functions)",
            "Closures – definition, practical uses, closure traps",
            "Closures – memoization, module pattern, IIFE",
            "Coding: write 3 closure exercises from scratch",
        ]
    },
    {
        "phase": "JavaScript", "week_label": "JS Week 2: this, Prototype, OOP",
        "subtopics": [
            "`this` keyword – all 4 rules (implicit, explicit, new, arrow)",
            "call / apply / bind – implement bind from scratch",
            "Prototype & Prototype Chain – draw the chain",
            "Class syntax vs Prototype – Object.create patterns",
            "Coding: implement inheritance two ways (class & prototype)",
        ]
    },
    {
        "phase": "JavaScript", "week_label": "JS Week 3: Async JavaScript",
        "subtopics": [
            "Event Loop – call stack, task queue, microtask queue (draw it)",
            "Promises – creation, chaining, Promise.all/race/allSettled",
            "async / await – patterns, error handling, parallel vs sequential",
            "Microtasks vs Macrotasks – setTimeout vs queueMicrotask",
            "Coding: implement Promise.all and Promise.race from scratch",
        ]
    },
    {
        "phase": "JavaScript", "week_label": "JS Week 4: Advanced JS & Browser APIs",
        "subtopics": [
            "Event Delegation, Bubbling, Capturing – implement custom event system",
            "Debouncing & Throttling – implement both from scratch",
            "Web Workers – Dedicated, Shared, Service Workers + postMessage + SharedArrayBuffer",
            "WeakMap/WeakRef, Generators, Modules (ESM vs CJS), Memory Leaks",
            "Coding: debounce + throttle + Web Worker for heavy computation from scratch",
        ]
    },

    # ── PHASE 2: TypeScript (Weeks 5–6) ──────────────────────────────────────
    {
        "phase": "TypeScript", "week_label": "TS Week 1: Types & Generics",
        "subtopics": [
            "Types vs Interfaces, Union, Intersection, Literal Types",
            "Generics – functions, classes, constraints, defaults",
            "Utility Types – Partial, Required, Pick, Omit, Record, Readonly",
            "Mapped Types – transform every key of a type",
            "Coding: build a type-safe API response wrapper with generics",
        ]
    },
    {
        "phase": "TypeScript", "week_label": "TS Week 2: Advanced Types",
        "subtopics": [
            "Conditional Types – T extends U ? X : Y, infer keyword",
            "Template Literal Types, `satisfies` operator",
            "Type Narrowing – typeof, instanceof, discriminated unions",
            "Decorators, ReturnType, Parameters, advanced inference",
            "Coding: build a type-safe form validation library",
        ]
    },

    # ── PHASE 3: Angular (Weeks 7–12) ────────────────────────────────────────
    {
        "phase": "Angular", "week_label": "Angular Week 1: Core Architecture",
        "subtopics": [
            "Angular architecture – modules, components, services overview",
            "Component Lifecycle Hooks – all 8 hooks with use cases",
            "Dependency Injection – providers, InjectionToken, hierarchical DI",
            "`inject()` function vs constructor DI – when to use which",
            "Coding: build a component with all lifecycle hooks logged",
        ]
    },
    {
        "phase": "Angular", "week_label": "Angular Week 2: Standalone & New APIs",
        "subtopics": [
            "Standalone Components – migration from NgModule pattern",
            "New control flow – @if, @for, @switch (vs old *ngIf, *ngFor)",
            "Deferrable Views – @defer, @placeholder, @loading, @error",
            "Signals – intro, signal(), computed(), effect()",
            "Coding: refactor an NgModule app to fully standalone",
        ]
    },
    {
        "phase": "Angular", "week_label": "Angular Week 3: Change Detection & Performance",
        "subtopics": [
            "Change Detection – Default strategy deep dive",
            "Change Detection – OnPush strategy, ChangeDetectorRef",
            "Signals – toSignal(), toObservable(), signal-based components",
            "Performance – trackBy, virtual scroll, lazy images",
            "Coding: optimize a slow list with OnPush + trackBy",
        ]
    },
    {
        "phase": "Angular", "week_label": "Angular Week 4: RxJS",
        "subtopics": [
            "RxJS – Observable creation, cold vs hot, Subject types",
            "RxJS – switchMap vs mergeMap vs concatMap vs exhaustMap",
            "RxJS – BehaviorSubject, ReplaySubject, combineLatest, forkJoin",
            "RxJS – takeUntilDestroyed, marble testing, custom operators",
            "Coding: build a type-ahead search with RxJS operators",
        ]
    },
    {
        "phase": "Angular", "week_label": "Angular Week 5: NgRx & Routing",
        "subtopics": [
            "NgRx – Store, Actions, Reducers, createFeature",
            "NgRx – Effects, createEffect, ofType, error handling",
            "NgRx – Selectors, createSelector, memoization",
            "Routing – Guards (CanActivate, CanDeactivate), Interceptors, Lazy Loading",
            "Coding: implement a full NgRx feature slice with effects",
        ]
    },
    {
        "phase": "Angular", "week_label": "Angular Week 6: SSR, Testing & Micro Frontends",
        "subtopics": [
            "SSR with Angular Universal – setup, TransferState",
            "Hydration – full & partial hydration",
            "Angular Testing – unit tests (Jasmine/Karma), TestBed, async",
            "Angular Testing Library, E2E basics (Playwright)",
            "Micro Frontends – Module Federation with Angular",
        ]
    },

    # ── PHASE 4: Frontend System Design (Weeks 13–16) ────────────────────────
    {
        "phase": "System Design", "week_label": "SD Week 1: Fundamentals",
        "subtopics": [
            "Browser rendering pipeline, Critical Rendering Path, Reflow/Repaint",
            "Frontend performance – Core Web Vitals, lazy loading, code splitting",
            "Design Systems & Component Libraries – tokens, variants, theming",
            "Component API design – props, events, slots, composition patterns",
            "Coding: design a themeable Button component system",
        ]
    },
    {
        "phase": "System Design", "week_label": "SD Week 2: Applications",
        "subtopics": [
            "Design an E-commerce Frontend – product listing, cart, checkout",
            "Design a Chat Application – WebSockets, presence, message sync",
            "Design YouTube Frontend – video player, recommendations, feeds",
            "Design Google Docs – collaborative editing, OT/CRDT basics",
            "Practice: 45-min whiteboard for any of the above",
        ]
    },
    {
        "phase": "System Design", "week_label": "SD Week 3: Auth, APIs & Scale",
        "subtopics": [
            "Authentication design – OAuth 2.0, JWT flow, refresh tokens",
            "Dashboard design – real-time data, polling vs WebSockets vs SSE",
            "File Upload system – chunked upload, progress, retry",
            "Notification system – push, in-app, email coordination",
            "Practice: 45-min design for File Upload or Notification system",
        ]
    },
    {
        "phase": "System Design", "week_label": "SD Week 4: Advanced Patterns",
        "subtopics": [
            "Micro Frontends – Module Federation, routing, shared state",
            "Autocomplete / Typeahead – debounce, caching, accessibility",
            "Rate Limiter (frontend) – token bucket, UX patterns",
            "Design Token System – multi-brand theming at scale",
            "Practice: 45-min design for Micro Frontends or Autocomplete",
        ]
    },
]

# Continue PHASE_WEEKS list
PHASE_WEEKS += [
    # ── PHASE 5: Backend + HTTP + Security (Weeks 17–21) ─────────────────────
    {
        "phase": "Backend", "week_label": "Backend Week 1: Node.js & Express",
        "subtopics": [
            "Node.js – event loop, streams, clustering, worker threads",
            "ExpressJS – middleware chain, routing, error handling patterns",
            "REST API design – versioning, HATEOAS, status codes, pagination",
            "Middleware patterns – logging, rate limiting, validation",
            "Coding: build a REST API with Express + error middleware",
        ]
    },
    {
        "phase": "Backend", "week_label": "Backend Week 2: HTTP Deep Dive",
        "subtopics": [
            "HTTP/1.1 – request/response cycle, headers, methods, status codes",
            "HTTPS – TLS handshake, certificates, mixed content",
            "HTTP/2 – multiplexing, header compression, server push",
            "HTTP/3 – QUIC protocol, 0-RTT, why it matters",
            "Caching – Cache-Control, ETag, Last-Modified, CDN caching layers",
        ]
    },
    {
        "phase": "Backend", "week_label": "Backend Week 3: Security",
        "subtopics": [
            "CORS – preflight, allowed origins, credentials – implement properly",
            "CSP (Content Security Policy) – directives, nonce, strict-dynamic",
            "XSS – reflected, stored, DOM-based – prevention & sanitization",
            "CSRF – SameSite cookies, CSRF tokens, double-submit pattern",
            "JWT attacks – alg:none, weak secrets, expiry bypass | OAuth attacks – redirect_uri, PKCE",
        ]
    },
    {
        "phase": "Backend", "week_label": "Backend Week 4: Databases & Auth",
        "subtopics": [
            "JWT – structure (header.payload.sig), signing, refresh token rotation",
            "OAuth 2.0 – authorization code + PKCE flow end-to-end",
            "SQL – joins, indexes, query optimization, EXPLAIN",
            "MongoDB – aggregation pipeline, indexes, atlas search",
            "Redis – caching patterns, pub/sub, session storage",
        ]
    },
    {
        "phase": "Backend", "week_label": "Backend Week 5: Infrastructure",
        "subtopics": [
            "Docker – Dockerfile, multi-stage builds, docker-compose",
            "Spring Boot basics – controllers, services, JPA, REST",
            "WebSockets – ws protocol, rooms, scaling with Redis adapter",
            "GraphQL basics – schema, resolvers, N+1 problem",
            "CDN & Compression – Brotli/Gzip, CDN edge caching, cache invalidation",
        ]
    },

    # ── PHASE 6: AI Engineering (Weeks 22–26) ────────────────────────────────
    {
        "phase": "AI Engineering", "week_label": "AI Week 1: LLM Fundamentals",
        "subtopics": [
            "LLM fundamentals – tokens, temperature, top-p, context window",
            "Prompt Engineering – zero-shot, few-shot, CoT, ReAct prompting",
            "OpenAI API – chat completions, streaming, function calling",
            "Claude API – tool use, system prompts, caching",
            "Coding: build a multi-turn chat with function calling",
        ]
    },
    {
        "phase": "AI Engineering", "week_label": "AI Week 2: RAG & Embeddings",
        "subtopics": [
            "Embeddings – what they are, cosine similarity, use cases",
            "Vector Databases – Chroma, Pinecone, pgvector – setup & query",
            "RAG – chunking strategies, retrieval, re-ranking",
            "RAG – evaluation (context recall, answer relevance)",
            "Coding: build a RAG pipeline over a PDF document",
        ]
    },
    {
        "phase": "AI Engineering", "week_label": "AI Week 3: Agents & Tools",
        "subtopics": [
            "AI Agents – ReAct pattern, tool calling, planning loops",
            "LangChain – chains, agents, memory, LCEL",
            "LangGraph – stateful agentic workflows, human-in-the-loop",
            "MCP (Model Context Protocol) – architecture, tools, resources",
            "Coding: build a multi-tool agent (web search + code executor)",
        ]
    },
    {
        "phase": "AI Engineering", "week_label": "AI Week 4: Production AI",
        "subtopics": [
            "Streaming responses – SSE, chunked transfer, frontend handling",
            "Ollama – local model serving, API, model management",
            "Amazon Q / Bedrock – integration patterns, enterprise use",
            "AI safety – guardrails, content filtering, hallucination mitigation",
            "Coding: add streaming to convopro_chatgpt app",
        ]
    },
    {
        "phase": "AI Engineering", "week_label": "AI Week 5: AI Engineering Review",
        "subtopics": [
            "AI evaluation frameworks – evals, benchmarks, automated testing",
            "LlamaIndex – document loading, indexing, query engines",
            "AI system design – latency, cost, caching, fallbacks",
            "Review all AI topics + prepare Boeing AI story for interviews",
            "Coding: polish convopro_chatgpt – final features + deploy",
        ]
    },

    # ── PHASE 7: Mock Loops + Revision (Weeks 27–30) ─────────────────────────
    {
        "phase": "Revision & Apply", "week_label": "Rev Week 1: Frontend Consolidation",
        "subtopics": [
            "Revise: JavaScript top 20 interview questions (write answers)",
            "Revise: Angular top 20 interview questions (write answers)",
            "Revise: TypeScript advanced patterns",
            "Revise: RxJS operators – code each from memory",
            "Mock coding round: 2 medium problems timed",
        ]
    },
    {
        "phase": "Revision & Apply", "week_label": "Rev Week 2: System Design Consolidation",
        "subtopics": [
            "Revise: all System Design patterns + draw diagrams",
            "Revise: Backend – HTTP, Security, REST",
            "Revise: all DSA patterns – write cheatsheet",
            "Revise: AI Engineering talking points for interviews",
            "Mock: Full system design round (60 min)",
        ]
    },
    {
        "phase": "Revision & Apply", "week_label": "Rev Week 3: Full Mock Loop 1",
        "subtopics": [
            "Full mock: JavaScript advanced coding (45 min)",
            "Full mock: Angular + TypeScript Q&A (45 min)",
            "Full mock: System Design round (60 min)",
            "Full mock: Behavioral round (45 min)",
            "Debrief: identify weak areas, update revision plan",
        ]
    },
    {
        "phase": "Revision & Apply", "week_label": "Rev Week 4: Final Prep & Apply",
        "subtopics": [
            "Weak area revision – targeted 3-hour session",
            "Company-specific prep: Atlassian / Adobe / Flipkart",
            "Company-specific prep: Razorpay / Salesforce / LinkedIn",
            "Full mock loop simulation (full day)",
            "Final review: personal pitch, salary negotiation, questions to ask",
        ]
    },
]

# ─────────────────────────────────────────────
# MASTER ROADMAP (all lessons in order)
# ─────────────────────────────────────────────

MASTER_ROADMAP = []
for pw in PHASE_WEEKS:
    for sub in pw["subtopics"]:
        MASTER_ROADMAP.append({
            "phase": pw["phase"],
            "week_label": pw["week_label"],
            "topic": sub,
            "status": "",
        })

# ─────────────────────────────────────────────
# REVISION TRACKER TOPICS (all core topics)
# ─────────────────────────────────────────────

REVISION_TOPICS = [
    # JavaScript
    "Execution Context & Call Stack",
    "Scope Chain & Closures",
    "this – 4 binding rules",
    "Prototype Chain",
    "Event Loop (draw it)",
    "Promises & async/await",
    "Debounce & Throttle (implement)",
    "Web Workers – postMessage, Transferable, SharedArrayBuffer",
    "Polyfills (implement)",
    # TypeScript
    "Generics & Utility Types",
    "Mapped & Conditional Types",
    "Type Narrowing & Discriminated Unions",
    # Angular
    "Lifecycle Hooks",
    "Change Detection (OnPush)",
    "Signals (computed, effect, toSignal)",
    "Standalone Components",
    "RxJS operators (switchMap/mergeMap/concatMap)",
    "NgRx (Store/Actions/Reducers/Effects)",
    "Guards & Interceptors",
    # System Design
    "Browser Rendering Pipeline",
    "Design: E-commerce Frontend",
    "Design: Chat Application",
    "Design: Authentication (OAuth/JWT)",
    "Design: Micro Frontends",
    # Backend
    "HTTP/1.1 vs HTTP/2 vs HTTP/3",
    "HTTPS & TLS Handshake",
    "CDN & Caching strategies",
    "CORS (implement correctly)",
    "CSP directives",
    "XSS prevention",
    "CSRF protection",
    "JWT attacks & mitigations",
    "OAuth attacks & PKCE",
    "REST API design",
    "Docker (build & run)",
    # AI
    "RAG pipeline",
    "Embeddings & Vector DB",
    "AI Agents (ReAct)",
    "MCP protocol",
    "LangGraph stateful workflow",
    # DSA
    "Sliding Window pattern",
    "Two Pointer pattern",
    "BFS/DFS (implement)",
    "Binary Search (implement)",
    "DP (coin change, knapsack)",
]

# ─────────────────────────────────────────────
# BEHAVIORAL STORIES
# ─────────────────────────────────────────────

BEHAVIORAL_STORIES = [
    ("AngularJS → Angular 12 migration (25 engineers)", "✅", 0),
    ("Leading a team of 25 engineers", "", 0),
    ("TypeScript standards adoption across teams", "", 0),
    ("AI workflow automation at Boeing (MCP + Claude)", "✅", 0),
    ("Mentoring junior engineers", "", 0),
    ("Handling a major technical disagreement", "", 0),
    ("Production incident / outage", "", 0),
    ("Code review process & culture building", "", 0),
    ("Work estimation & planning accuracy", "", 0),
    ("Technical decision you influenced significantly", "", 0),
    ("Why leaving Boeing? (rehearsed answer)", "", 0),
    ("60-second personal pitch", "", 0),
    ("Salary negotiation talking points", "", 0),
]

# ─────────────────────────────────────────────
# SCHEDULE BUILDER
# ─────────────────────────────────────────────

def build_daily_schedule():
    """
    One topic per day — no multi-topic cramming.

    Mon – Core subtopic 1
    Tue – Core subtopic 2
    Wed – Core subtopic 3
    Thu – Core subtopic 4
    Fri – Core subtopic 5  (coding / practice session)
    Sat – Mock Interview + Behavioral + Revision  (dedicated block, no theory)
    Sun – Rest + Project + Portfolio  (light, protected)
    """
    rows = []
    current = START_DATE
    phase_week_idx = 0
    mock_i = sun_i = 0

    for week_offset in range(TOTAL_WEEKS):
        week_num = week_offset + 1
        pw = PHASE_WEEKS[phase_week_idx % len(PHASE_WEEKS)]
        phase_week_idx += 1
        subs = pw["subtopics"]  # exactly 5 subtopics — one per weekday

        for day_offset in range(7):
            d = current + timedelta(days=day_offset)
            dow = d.weekday()
            month_num = min(((d - START_DATE).days // 28) + 1, 7)

            if dow == 0:    # Monday
                full = subs[0]
                focus = pw["phase"]
                planned_hrs = 2

            elif dow == 1:  # Tuesday
                full = subs[1]
                focus = pw["phase"]
                planned_hrs = 2

            elif dow == 2:  # Wednesday
                full = subs[2]
                focus = pw["phase"]
                planned_hrs = 2

            elif dow == 3:  # Thursday
                full = subs[3]
                focus = pw["phase"]
                planned_hrs = 2

            elif dow == 4:  # Friday
                full = subs[4]
                focus = pw["phase"]
                planned_hrs = 2

            elif dow == 5:  # Saturday — Mock + Revision only
                # Mock topic matches what was studied this week
                mock = f"Q&A + coding: {pw['week_label'].split(': ', 1)[-1]}"
                # Behavioral rotates through stories you already know (Boeing experience)
                beh  = BEHAVIORAL_TOPICS[mock_i % len(BEHAVIORAL_TOPICS)]
                mock_i += 1
                full  = f"Mock: {mock}  |  Behavioral: {beh}  |  Revise weak areas from this week"
                focus = "Mock + Revision"
                planned_hrs = 3

            else:           # Sunday — rest / project / portfolio
                sun = SUNDAY_TOPICS[sun_i % len(SUNDAY_TOPICS)]
                sun_i += 1
                full  = sun
                focus = "Rest + Project"
                planned_hrs = 1

            rows.append({
                "week":        week_num,
                "month":       f"Month {month_num}",
                "phase":       pw["phase"],
                "week_label":  pw["week_label"],
                "date_str":    d.strftime("%d %b %Y"),
                "day":         d.strftime("%A"),
                "focus":       focus,
                "full_topic":  full,
                "planned_hrs": planned_hrs,
                "actual_hrs":  "",
                "done":        "",
                "coding_done": "",
                "notes":       "",
            })

        current += timedelta(weeks=1)

    return rows

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

PHASE_COLORS = {
    "JavaScript":       "FFF2CC",
    "TypeScript":       "E2EFDA",
    "Angular":          "D9EAD3",
    "System Design":    "CFE2F3",
    "Backend":          "D9D2E9",
    "AI Engineering":   "EAD1DC",
    "Revision & Apply": "FCE5CD",
    "Mock + Revision + Project":  "FFE0B2",
    "Reading + Portfolio + Rest": "E8F5E9",
}

HDR_COLOR    = "1F4E79"
PHASE_HDR    = "BDD7EE"
SHEET2_HDR   = "1F4E79"
ACCENT       = "2E74B5"
LIGHT_GRAY   = "F2F2F2"

THIN   = Side(style="thin",   color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def mf(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def hdr_cell(ws, row, col, value, bg=HDR_COLOR, fg="FFFFFF", bold=True, size=10,
             halign="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size)
    c.fill      = mf(bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    c.border    = BORDER
    return c

def data_cell(ws, row, col, value, bg="FFFFFF", bold=False, size=9,
              halign="left", wrap=True, fg="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size)
    c.fill      = mf(bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    c.border    = BORDER
    return c

def add_dropdown(ws, col_letter, start_row, end_row, formula, title="", error=""):
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=True,
        showDropDown=False, showErrorMessage=bool(error),
        errorTitle=title, error=error,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)
    return dv

def freeze(ws, cell): ws.freeze_panes = cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ─────────────────────────────────────────────
# SHEET 1: DASHBOARD
# ─────────────────────────────────────────────

def write_dashboard(wb, rows):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="🎯  Interview Prep Tracker – Mriganka Sekhar Sarkar")
    c.font      = Font(bold=True, size=16, color="FFFFFF")
    c.fill      = mf(HDR_COLOR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 36)

    ws.merge_cells("A2:D2")
    c = ws.cell(row=2, column=1, value="Boeing → Product Company | 30-Week Interview Preparation Program")
    c.font      = Font(size=10, color="595959", italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 2, 20)

    # Stats
    stats = [
        ("Current Week",      "=MAX('📅 Daily Planner'!A:A)",             "week"),
        ("Topics Completed",  "=COUNTIF('📅 Daily Planner'!K:K,\"Completed\")", "count"),
        ("Total Days",        len(rows),                                   "count"),
        ("Study Hrs (Planned)","=SUM('📅 Daily Planner'!I:I)",            "hrs"),
        ("Mock Interviews",   "=COUNTIF('📅 Daily Planner'!F:F,\"Mock + Revision + Project\")", "count"),
        ("Phase",             "=IFERROR(INDEX('📅 Daily Planner'!C:C,MATCH(MAX('📅 Daily Planner'!A:A),'📅 Daily Planner'!A:A,0)),\"–\")", "text"),
    ]

    labels_row  = 4
    values_row  = 5
    stat_colors = ["CFE2F3", "D9EAD3", "FFF2CC", "EAD1DC", "FCE5CD", "D9D2E9"]

    for i, (label, formula, _) in enumerate(stats):
        col = i + 1
        c = ws.cell(row=labels_row, column=col, value=label)
        c.font = Font(bold=True, size=9, color="595959")
        c.fill = mf(stat_colors[i])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        set_row_height(ws, labels_row, 18)

        v = ws.cell(row=values_row, column=col, value=formula)
        v.font = Font(bold=True, size=18, color=HDR_COLOR)
        v.fill = mf(stat_colors[i])
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.border = BORDER
        set_row_height(ws, values_row, 40)
        set_col_width(ws, col, 22)

    # Phase roadmap summary
    phase_row = 7
    ws.merge_cells(f"A{phase_row}:D{phase_row}")
    c = ws.cell(row=phase_row, column=1, value="PHASE ROADMAP")
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = mf(ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    set_row_height(ws, phase_row, 22)

    phases = [
        ("Phase 1", "Weeks  1–4",  "JavaScript",        "Rock-solid JS foundation"),
        ("Phase 2", "Weeks  5–6",  "TypeScript",         "Advanced types mastery"),
        ("Phase 3", "Weeks  7–12", "Angular",            "Deep Angular + RxJS + NgRx"),
        ("Phase 4", "Weeks 13–16", "Frontend Sys Design","All major FE design patterns"),
        ("Phase 5", "Weeks 17–21", "Backend + HTTP + Security", "Node, REST, JWT, CORS, XSS, OAuth"),
        ("Phase 6", "Weeks 22–26", "AI Engineering",    "RAG, Agents, MCP, LangGraph"),
        ("Phase 7", "Weeks 27–30", "Revision & Apply",  "Mock loops, apply, negotiate"),
    ]

    ph_colors = ["FFF2CC","E2EFDA","D9EAD3","CFE2F3","D9D2E9","EAD1DC","FCE5CD"]
    hdr_row = phase_row + 1
    for ci, h in enumerate(["Phase","Weeks","Focus","Goal"], 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = Font(bold=True, size=9, color=HDR_COLOR); c.fill = mf(LIGHT_GRAY)
        c.alignment = Alignment(horizontal="center"); c.border = BORDER
    set_row_height(ws, hdr_row, 18)

    for pi, (ph, wks, focus, goal) in enumerate(phases):
        r = hdr_row + 1 + pi
        bg = ph_colors[pi]
        for ci, val in enumerate([ph, wks, focus, goal], 1):
            data_cell(ws, r, ci, val, bg=bg, size=9,
                      halign="center" if ci <= 2 else "left")
        set_row_height(ws, r, 18)

    # Tips
    tip_row = hdr_row + len(phases) + 3
    ws.merge_cells(f"A{tip_row}:D{tip_row}")
    c = ws.cell(row=tip_row, column=1, value="💡 KEY TIPS")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = mf(ACCENT)
    c.alignment = Alignment(horizontal="left")

    tips = [
        "• Mon–Fri: Same core topic all week. Different sub-topics each day. Deep learning, not scatter.",
        "• Wed always adds DSA. Thu always adds AI/System Design. Fri always adds LeetCode + Behavioral.",
        "• Saturday = Mock Interview (NON-NEGOTIABLE). This is where real interview skills are built.",
        "• Sunday = Rest + Project + Portfolio. Protect this day from overwork.",
        "• Use Sheet 4 (Revision Tracker) after each topic. Spaced repetition beats re-reading.",
        "• Biggest differentiator: AI Engineering (MCP, LangGraph, Claude at Boeing). Lead with this.",
        "• Start applying from Week 17 (Month 4). Real interviews teach more than mock prep.",
        "• Security topics (CORS, XSS, CSRF, JWT/OAuth attacks) are asked in senior interviews. Know them well.",
    ]
    for ti, tip in enumerate(tips):
        r = tip_row + 1 + ti
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=tip)
        c.font = Font(size=9); c.fill = mf("FAFAFA")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        set_row_height(ws, r, 16)

# ─────────────────────────────────────────────
# SHEET 2: MASTER ROADMAP
# ─────────────────────────────────────────────

def write_master_roadmap(wb):
    ws = wb.create_sheet("🗺 Master Roadmap")
    ws.sheet_view.showGridLines = False

    headers    = ["Phase", "Week Cluster", "Lesson / Topic", "Status"]
    col_widths = [18, 32, 62, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws, 1, ci, h); set_col_width(ws, ci, w)
    set_row_height(ws, 1, 22)
    freeze(ws, "A2")

    add_dropdown(ws, "D", 2, len(MASTER_ROADMAP) + 5,
                 '"✅ Done,🔄 In Progress,⬜ Not Started"',
                 "Status", 'Choose: ✅ Done, 🔄 In Progress, or ⬜ Not Started')

    ph_colors = {
        "JavaScript":       "FFF2CC",
        "TypeScript":       "E2EFDA",
        "Angular":          "D9EAD3",
        "System Design":    "CFE2F3",
        "Backend":          "D9D2E9",
        "AI Engineering":   "EAD1DC",
        "Revision & Apply": "FCE5CD",
    }

    current_wl = None
    for ri, lesson in enumerate(MASTER_ROADMAP, 2):
        bg = ph_colors.get(lesson["phase"], "FFFFFF")
        # Slightly darken week cluster header rows
        is_new_cluster = (lesson["week_label"] != current_wl)
        if is_new_cluster:
            current_wl = lesson["week_label"]
            row_bg = "E8E8E8"
            for ci, val in enumerate([lesson["phase"], lesson["week_label"], "", ""], 1):
                c = ws.cell(row=ri, column=ci, value=val if ci <= 2 else "")
                c.font = Font(bold=True, size=9, color=HDR_COLOR)
                c.fill = mf(row_bg); c.border = BORDER
                c.alignment = Alignment(horizontal="left", vertical="center")
            set_row_height(ws, ri, 16)
        else:
            data_cell(ws, ri, 1, lesson["phase"], bg=bg, size=9)
            data_cell(ws, ri, 2, lesson["week_label"], bg=bg, size=9)
            data_cell(ws, ri, 3, lesson["topic"], bg=bg, size=9)
            data_cell(ws, ri, 4, "⬜ Not Started", bg=bg, size=9, halign="center")
            set_row_height(ws, ri, 15)

# ─────────────────────────────────────────────
# SHEET 3: DAILY PLANNER (main tracker)
# ─────────────────────────────────────────────

def write_daily_planner(wb, rows):
    ws = wb.create_sheet("📅 Daily Planner")
    ws.sheet_view.showGridLines = False

    headers    = ["Week","Month","Phase","Week Focus","Date","Day",
                  "Focus Area","Today's Topic / Task",
                  "Planned Hrs","Actual Hrs","Done","Coding Done","Notes"]
    col_widths = [6, 10, 14, 28, 14, 12, 26, 70, 10, 10, 14, 12, 28]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws, 1, ci, h, wrap=True); set_col_width(ws, ci, w)
    set_row_height(ws, 1, 30)
    freeze(ws, "A2")

    total_data_rows = len(rows)
    add_dropdown(ws, "K", 2, total_data_rows + 10,
                 '"Completed,In Progress"', "Status",
                 'Select "Completed" or "In Progress"')
    add_dropdown(ws, "L", 2, total_data_rows + 10,
                 '"Yes,No"', "Coding", 'Select Yes or No')

    current_phase = None
    excel_row = 2

    for r in rows:
        # Phase separator
        if r["phase"] != current_phase:
            current_phase = r["phase"]
            bg_sep = PHASE_COLORS.get(current_phase, PHASE_HDR)
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=len(headers))
            c = ws.cell(row=excel_row, column=1,
                        value=f"━━━  {current_phase.upper()}  ━━━")
            c.font = Font(bold=True, size=10, color=HDR_COLOR)
            c.fill = mf(bg_sep)
            c.alignment = Alignment(horizontal="left", vertical="center")
            set_row_height(ws, excel_row, 18)
            excel_row += 1

        # Determine row color
        focus = r["focus"]
        if "Mock" in focus or "Revision" in focus:
            bg = PHASE_COLORS["Mock + Revision + Project"]
        elif "Reading" in focus or "Portfolio" in focus:
            bg = PHASE_COLORS["Reading + Portfolio + Rest"]
        else:
            bg = PHASE_COLORS.get(r["phase"], "FFFFFF")

        vals = [
            r["week"], r["month"], r["phase"], r["week_label"],
            r["date_str"], r["day"], r["focus"], r["full_topic"],
            r["planned_hrs"], r["actual_hrs"], r["done"],
            r["coding_done"], r["notes"],
        ]

        for ci, val in enumerate(vals, 1):
            bold = ci in (4, 8)
            halign = "center" if ci in (1,2,5,6,9,10,11,12) else "left"
            data_cell(ws, excel_row, ci, val, bg=bg, bold=bold,
                      size=9, halign=halign, wrap=True)

        set_row_height(ws, excel_row, 30)
        excel_row += 1

# ─────────────────────────────────────────────
# SHEET 4: REVISION TRACKER (spaced repetition)
# ─────────────────────────────────────────────

def write_revision_tracker(wb):
    ws = wb.create_sheet("🔁 Revision Tracker")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="🔁 Spaced Repetition Tracker — Review each topic at 1 day, 7 days, 30 days, then mark Mastered")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = mf(HDR_COLOR)
    c.alignment = Alignment(horizontal="left", vertical="center")
    set_row_height(ws, 1, 22)

    headers    = ["Topic", "Category", "1 Day ✓", "7 Days ✓", "30 Days ✓", "Mastered ✓"]
    col_widths = [42, 16, 10, 10, 10, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws, 2, ci, h, size=9); set_col_width(ws, ci, w)
    set_row_height(ws, 2, 20)
    freeze(ws, "A3")

    check_cols = ["C", "D", "E", "F"]
    total = len(REVISION_TOPICS) + 5
    for col_letter in check_cols:
        add_dropdown(ws, col_letter, 3, total,
                     '"✅,⬜"', "Revision", 'Select ✅ when reviewed')

    category_map = {
        "Execution Context & Call Stack": "JavaScript",
        "Scope Chain & Closures": "JavaScript",
        "this – 4 binding rules": "JavaScript",
        "Prototype Chain": "JavaScript",
        "Event Loop (draw it)": "JavaScript",
        "Promises & async/await": "JavaScript",
        "Debounce & Throttle (implement)": "JavaScript",
        "Polyfills (implement)": "JavaScript",
        "Generics & Utility Types": "TypeScript",
        "Mapped & Conditional Types": "TypeScript",
        "Type Narrowing & Discriminated Unions": "TypeScript",
        "Lifecycle Hooks": "Angular",
        "Change Detection (OnPush)": "Angular",
        "Signals (computed, effect, toSignal)": "Angular",
        "Standalone Components": "Angular",
        "RxJS operators (switchMap/mergeMap/concatMap)": "Angular",
        "NgRx (Store/Actions/Reducers/Effects)": "Angular",
        "Guards & Interceptors": "Angular",
        "Browser Rendering Pipeline": "System Design",
        "Design: E-commerce Frontend": "System Design",
        "Design: Chat Application": "System Design",
        "Design: Authentication (OAuth/JWT)": "System Design",
        "Design: Micro Frontends": "System Design",
        "HTTP/1.1 vs HTTP/2 vs HTTP/3": "Backend",
        "HTTPS & TLS Handshake": "Backend",
        "CDN & Caching strategies": "Backend",
        "CORS (implement correctly)": "Security",
        "CSP directives": "Security",
        "XSS prevention": "Security",
        "CSRF protection": "Security",
        "JWT attacks & mitigations": "Security",
        "OAuth attacks & PKCE": "Security",
        "REST API design": "Backend",
        "Docker (build & run)": "Backend",
        "RAG pipeline": "AI Engineering",
        "Embeddings & Vector DB": "AI Engineering",
        "AI Agents (ReAct)": "AI Engineering",
        "MCP protocol": "AI Engineering",
        "LangGraph stateful workflow": "AI Engineering",
        "Sliding Window pattern": "DSA",
        "Two Pointer pattern": "DSA",
        "BFS/DFS (implement)": "DSA",
        "Binary Search (implement)": "DSA",
        "DP (coin change, knapsack)": "DSA",
    }

    cat_colors = {
        "JavaScript": "FFF2CC", "TypeScript": "E2EFDA", "Angular": "D9EAD3",
        "System Design": "CFE2F3", "Backend": "D9D2E9", "Security": "F4CCCC",
        "AI Engineering": "EAD1DC", "DSA": "FCE5CD",
    }

    for ri, topic in enumerate(REVISION_TOPICS, 3):
        cat = category_map.get(topic, "")
        bg  = cat_colors.get(cat, "FFFFFF")
        data_cell(ws, ri, 1, topic, bg=bg, size=9)
        data_cell(ws, ri, 2, cat, bg=bg, size=9, halign="center")
        for ci in range(3, 7):
            data_cell(ws, ri, ci, "⬜", bg=bg, size=10, halign="center")
        set_row_height(ws, ri, 16)

# ─────────────────────────────────────────────
# SHEET 5: MOCK INTERVIEW LOG
# ─────────────────────────────────────────────

def write_mock_log(wb):
    ws = wb.create_sheet("🎤 Mock Interviews")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value="🎤 Mock Interview Log — Track every mock, score it, note weaknesses")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = mf(HDR_COLOR)
    c.alignment = Alignment(horizontal="left", vertical="center")
    set_row_height(ws, 1, 22)

    headers    = ["Date", "Company / Interviewer", "Topic / Round", "Score (1–10)", "Weakness Identified", "Action Item"]
    col_widths = [14, 24, 26, 14, 32, 32]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws, 2, ci, h, size=9); set_col_width(ws, ci, w)
    set_row_height(ws, 2, 20)
    freeze(ws, "A3")

    # Pre-fill example rows
    examples = [
        ("21 Jul 2026", "Self / Peer", "JS – Closures & Event Loop", "7/10", "Microtask queue ordering", "Re-read & draw the event loop 3 times"),
        ("28 Jul 2026", "Self / Peer", "Angular – Change Detection", "",     "",                          ""),
        ("04 Aug 2026", "Self / Peer", "System Design – Chat App",   "",     "",                          ""),
    ]
    for ri, row in enumerate(examples, 3):
        bg = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            data_cell(ws, ri, ci, val, bg=bg, size=9)
        set_row_height(ws, ri, 18)

    # Empty rows for user to fill in
    for ri in range(len(examples) + 3, 53):
        bg = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
        for ci in range(1, 7):
            data_cell(ws, ri, ci, "", bg=bg, size=9)
        set_row_height(ws, ri, 18)


# ─────────────────────────────────────────────
# SHEET 6: BEHAVIORAL STORIES
# ─────────────────────────────────────────────

def write_behavioral_stories(wb):
    ws = wb.create_sheet("💬 Behavioral Stories")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1,
                value="💬 STAR Story Tracker — Prepare, practice, and track each behavioral story")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = mf(HDR_COLOR)
    c.alignment = Alignment(horizontal="left", vertical="center")
    set_row_height(ws, 1, 22)

    headers    = ["STAR Story", "Ready?", "Times Practiced", "Last Practiced", "Notes / Improvements"]
    col_widths = [44, 12, 18, 16, 40]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws, 2, ci, h, size=9); set_col_width(ws, ci, w)
    set_row_height(ws, 2, 20)

    add_dropdown(ws, "B", 3, len(BEHAVIORAL_STORIES) + 5,
                 '"✅ Ready,🔄 In Progress,⬜ Not Started"',
                 "Ready", 'Choose story readiness')

    row_colors = ["FFF2CC","E2EFDA","D9EAD3","CFE2F3","D9D2E9","EAD1DC","FCE5CD",
                  "FFF2CC","E2EFDA","D9EAD3","CFE2F3","D9D2E9","EAD1DC"]

    for ri, (story, ready, practiced) in enumerate(BEHAVIORAL_STORIES, 3):
        bg = row_colors[(ri - 3) % len(row_colors)]
        ready_val = "✅ Ready" if ready == "✅" else "⬜ Not Started"
        data_cell(ws, ri, 1, story,       bg=bg, size=9)
        data_cell(ws, ri, 2, ready_val,   bg=bg, size=9, halign="center")
        data_cell(ws, ri, 3, practiced,   bg=bg, size=9, halign="center")
        data_cell(ws, ri, 4, "",          bg=bg, size=9, halign="center")
        data_cell(ws, ri, 5, "",          bg=bg, size=9)
        set_row_height(ws, ri, 18)

    # Tips for behavioral section
    tip_row = len(BEHAVIORAL_STORIES) + 5
    tips = [
        "💡 STAR FORMAT REMINDER",
        "S – Situation: Set the context (project, team size, timeline)",
        "T – Task: What was your specific responsibility?",
        "A – Action: Exactly what YOU did (use 'I', not 'we')",
        "R – Result: Measurable outcome (% improvement, engineers impacted, time saved)",
        "",
        "🔑 BOEING AI STORY: Lead every interview with this.",
        "   'I designed an automated defect workflow using Claude MCP + Word + Git → auto work items, auto fixes, auto PRs'",
        "   Quantify: 'Saved X hours/week', 'Reduced defect resolution time by Y%'",
    ]
    for ti, tip in enumerate(tips):
        r = tip_row + ti
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=tip)
        if tip.startswith(("💡", "🔑")):
            c.font = Font(bold=True, size=10, color=HDR_COLOR)
        else:
            c.font = Font(size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        set_row_height(ws, r, 16)

# ─────────────────────────────────────────────
# MAIN WRITER
# ─────────────────────────────────────────────

def write_excel(rows):
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_dashboard(wb, rows)
    write_master_roadmap(wb)
    write_daily_planner(wb, rows)
    write_revision_tracker(wb)
    write_mock_log(wb)
    write_behavioral_stories(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Workbook saved: {OUTPUT_FILE}")
    print(f"    Days generated : {len(rows)}")
    print(f"    Weeks          : {TOTAL_WEEKS}")
    print(f"    Sheets         : Dashboard | Master Roadmap | Daily Planner | Revision Tracker | Mock Interviews | Behavioral Stories")
    print(f"\n    Tip: Run adjust_schedule.py to handle gap days.\n")


if __name__ == "__main__":
    schedule = build_daily_schedule()
    write_excel(schedule)
